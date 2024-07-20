import requests
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color
from openpyxl.styles.colors import BLUE
from datetime import datetime
import re

# Function to read the GitHub access token from a file
def read_token_from_file(file_path):
    with open(file_path, 'r') as file:
        return file.read().strip()
token_file_path = "C:\\Users\\wquraishi\\Documents\\GitHub-Config\/github_token.txt"
access_token = read_token_from_file(token_file_path)

def sanitize_for_excel(text):
    if not text:
        return text  # or return "" to avoid NoneType issues
    # Remove HTML tags
    text = re.sub('<[^<]+?>', '', text)
    # Replace URLs with a simple placeholder or remove them
    text = re.sub(r'http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\\(\\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+', '[LINK]', text)
    # Remove or replace other illegal characters as needed
    text = text.replace('\n', ' ').replace('\r', '').replace('\t', ' ')
    # Truncate to avoid Excel cell character limit issues
    text = (text[:32767]) if len(text) > 32767 else text
    return text

def fetch_all_items(base_url, headers):
    items = []
    page = 1
    while True:
        # Construct the full URL with query parameters for each request
        api_url = f"{base_url}?state=all&page={page}&per_page=100"
        response = requests.get(api_url, headers=headers)
        print(f"Fetching {api_url}")  # Debug print to check the constructed URL
        
        if response.status_code == 200:
            data = response.json()
            if not data:
                break  # No more data, exit the loop
            items.extend(data)
            page += 1
        else:
            print(f"Failed to fetch data. Status Code: {response.status_code}. Response: {response.text}")
            break
    return items


repo_owner = "microsoft"
repo_name = "azurechat"

headers = {"Authorization": f"Bearer {access_token}"}

issues_url = f"https://api.github.com/repos/{repo_owner}/{repo_name}/issues"
pulls_url = f"https://api.github.com/repos/{repo_owner}/{repo_name}/pulls"

# Fetch issues and pull requests with pagination
issues_data = fetch_all_items(issues_url, headers)
pulls_data = fetch_all_items(pulls_url, headers)

# The rest of your script remains the same...

# Create an Excel workbook
wb = Workbook()
sheet = wb.active
sheet.title = "Issues and Pull Requests"
# Set header row font to bold...
# Continue as before.

# Write the header row
header_row = ["Number", "Type", "Title", "Body", "Assignees", "Labels", "Milestone", "State", "Reviewers", "Committers"]
for col_num, header in enumerate(header_row, 1):
    col_letter = get_column_letter(col_num)
    sheet[f"{col_letter}1"] = header

# Write issues data to the Excel file
for row_num, issue in enumerate(issues_data, 2):
    issue_number = issue["number"]
    issue_type = "Issue"
    issue_title = issue["title"]
    issue_body = issue["body"]
    issue_body = sanitize_for_excel(issue["body"])
    issue_assignees = ",".join(assignee["login"] for assignee in issue["assignees"])
    issue_labels = ",".join(label["name"] for label in issue["labels"])
    issue_milestone = issue["milestone"]["title"] if issue["milestone"] else ""
    issue_state = issue["state"]
    issue_url = issue["html_url"]

    cell = f"A{row_num}"
    sheet[cell].hyperlink = f'{issue_url}'
    sheet[cell].value = issue_number
    sheet[cell].font = Font(color=Color(rgb=BLUE))
    sheet[f"B{row_num}"] = issue_type
    sheet[f"C{row_num}"] = issue_title
    #Commented due to IllegalCharacterError
    #sheet[f"D{row_num}"] = issue_body
    sheet[f"E{row_num}"] = issue_assignees
    sheet[f"F{row_num}"] = issue_labels
    sheet[f"G{row_num}"] = issue_milestone
    sheet[f"H{row_num}"] = issue_state

# Write pull requests data to the Excel file
for pull_num, pull in enumerate(pulls_data, row_num + 1):
    pull_number = pull["number"]
    pull_type = "Pull Request"
    pull_title = pull["title"]
    #pull_body = pull["body"]
    pull_assignees = ",".join(assignee["login"] for assignee in pull["assignees"])
    pull_labels = ",".join(label["name"] for label in pull["labels"])
    pull_milestone = pull["milestone"]["title"] if pull["milestone"] else ""
    pull_state = pull["state"]
    pull_url = pull["html_url"]
    pull_reviewers = ",".join(reviewer["login"] for reviewer in pull["requested_reviewers"])
    pull_committers = pull["user"]["login"]

    cell = f"A{pull_num}"
    sheet[cell].hyperlink = f'{pull_url}'
    sheet[cell].value = pull_number
    sheet[cell].font = Font(color=Color(rgb=BLUE))
    sheet[f"B{pull_num}"] = pull_type
    sheet[f"C{pull_num}"] = pull_title
    #sheet[f"D{pull_num}"] = pull_body
    sheet[f"E{pull_num}"] = pull_assignees
    sheet[f"F{pull_num}"] = pull_labels
    sheet[f"G{pull_num}"] = pull_milestone
    sheet[f"H{pull_num}"] = pull_state
    sheet[f"I{pull_num}"] = pull_reviewers
    sheet[f"J{pull_num}"] = pull_committers

# Adjust column widths
for col in sheet.columns:
    max_length = 0
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    adjusted_width = (max_length + 2) * 1.2
    sheet.column_dimensions[col[0].column_letter].width = adjusted_width

# Save the workbook as an Excel file
if issues_data or pulls_data:
    print("Data fetched, writing to Excel...")
else:
    print("No data fetched, please check the fetch logic.")

# Generate output filename with current datetime suffix
current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")
output_filename = f"issues_and_pull_requests_{current_datetime}.xlsx"
wb.save(output_filename)
